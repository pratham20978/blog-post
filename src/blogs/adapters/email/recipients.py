"""Reading recipient addresses out of an uploaded spreadsheet.

Deliberately forgiving about *shape* and strict about *content*.

Forgiving, because the sheet comes from a person: it may have a header row or
none, the addresses may be in column A or column F, there may be several sheets,
and there is usually a stray note somewhere. Demanding a named column would mean
rejecting a file that plainly contains the addresses, so every cell is scanned
and anything that is an address is taken.

Strict, because the output is a send list. Addresses are lowercased and
deduplicated, and anything that does not match the same pattern ``EmailStr``
enforces is discarded rather than guessed at. A malformed address costs a
bounce, and enough bounces cost the sending domain its reputation.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

from openpyxl import load_workbook

from blogs.contracts.common import ErrorCategory
from blogs.core.errors import BlogPlatformError

#: The pattern `EmailStr` enforces, applied to a whole cell. Anchored on both
#: ends so "email: a@b.com" is rejected rather than half-parsed — a cell like
#: that means the sheet is not what we think it is.
_ADDRESS = re.compile(r"^[^@\s]+@[^@\s.]+(\.[^@\s.]+)+$")


@dataclass(frozen=True, slots=True)
class RecipientSheet:
    """What was found, and what was passed over."""

    addresses: tuple[str, ...]
    #: Cells that held something non-empty which was not an address — header
    #: labels and name columns included, so this is a rough signal rather than
    #: an error count. Reported back because a value far larger than the number
    #: of addresses usually means the wrong file, or addresses stored as images.
    skipped: int = 0
    #: Addresses that appeared more than once.
    duplicates: int = 0
    sheets: tuple[str, ...] = field(default=())


def read_recipients(data: bytes, *, limit: int) -> RecipientSheet:
    """Extract addresses from an ``.xlsx``.

    ``limit`` is a blast radius, not a performance guard: past it the whole
    request is refused rather than half-sent, because a truncated announcement
    is worse than none — you cannot tell who already received it.
    """
    try:
        # `read_only` streams rows instead of building the whole object model,
        # and `data_only` takes cached values rather than formula text.
        workbook = load_workbook(
            io.BytesIO(data), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:
        raise BlogPlatformError(
            ErrorCategory.REQUEST_INVALID,
            safe_message="That file could not be read as a spreadsheet.",
            safe_details={"reason": "NOT_XLSX"},
        ) from exc

    seen: dict[str, None] = {}
    skipped = 0
    duplicates = 0

    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    text = str(cell).strip()
                    if not text:
                        continue
                    if not _ADDRESS.match(text):
                        skipped += 1
                        continue
                    address = text.lower()
                    if address in seen:
                        duplicates += 1
                        continue
                    seen[address] = None
                    if len(seen) > limit:
                        raise BlogPlatformError(
                            ErrorCategory.REQUEST_INVALID,
                            safe_message=(
                                f"That sheet holds more than {limit} addresses. "
                                "Split it, or raise BLOGS_ANNOUNCE_MAX_RECIPIENTS."
                            ),
                            safe_details={"limit": limit},
                        )
        names = tuple(worksheet.title for worksheet in workbook.worksheets)
    finally:
        # Read-only workbooks hold an open zip handle until closed.
        workbook.close()

    return RecipientSheet(
        addresses=tuple(seen),
        skipped=skipped,
        duplicates=duplicates,
        sheets=names,
    )
